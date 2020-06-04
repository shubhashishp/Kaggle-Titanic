# -*- coding: utf-8 -*-
"""
Created on Mon Nov 18 12:33:11 2019

@author: shubhashish.p
"""

import pandas as pd
from sklearn.model_selection import train_test_split

path = "C:\\Shubhashish\\Experiments\\Kaggle\\Titanic\\titanic\\";
file = path+"train.csv"
df = pd.read_csv(file)

######### EDA
df.head()
df.shape
df.info()
df.isnull().sum()

#viszualization
import matplotlib.pyplot as plt 
%matplotlib inline
import seaborn as sns

sns.set()

def bar_chart(feature):
    survvived = df[df['Survived']==1][feature].value_counts()
    dead = df[df['Survived']==0][feature].value_counts()
    df_plot = pd.DataFrame([survvived,dead])
    df_plot.index = ["Survived","Dead"]
    df_plot.plot(kind='bar',stacked=True, figsize=(10,5))
    
bar_chart("Sex")
bar_chart("Pclass")
bar_chart("SibSp")
bar_chart("Parch")
bar_chart("Embarked")
sns.pairplot(data=df,hue="Survived",dropna="True")




############# 

df['Title'] = ""

#adding new feature



def adding_new_feature(df):   

    for index in range(len(df['Ticket'])):   
        t = df['Ticket'][index]
        n = df['Name'][index]
        
        if "" in t:
            x = t.split(" ")
            parts = len(x)
            number = x[parts-1]
            rem = t.replace(number,"")
            rem = rem.replace("/","")
            rem = rem.replace(".","")
            if number== "LINE":
                number=0  
#            df['T_number'][index] = int(number)
#            df['T_group'][index] = rem.strip().lower()
#        else:
#            df['T_number'][index] = int(df['T_number'][index])
        
        ns = n.split(", ")
        nt = ns[1].split(" ")
        if ( nt[0].strip() == "Mr."):
            df['Title'][index] = 0
        elif ( nt[0].strip() == "Miss."):
            df['Title'][index] = 1
        elif ( nt[0].strip() == "Mrs."):
            df['Title'][index] = 2
        elif ( nt[0].strip() == "Master."):
            df['Title'][index] = 3
        else:
            df['Title'][index] = 4
        
        
                 
                
adding_new_feature(df)


sns.heatmap(df.corr(), annot=True)

list(df.columns) # for column names
df.info() # shows total data 
bar_chart('Title')
# find columns and do the traetment of missing values
# by df.count we found that column "Age"(714/891),"Cabin"(204/891),"Embark"(889/891) are have some missing values 

def do_preprocessing(df):
    
    df["Cabin"].fillna(0,inplace=True)
    df['Cabin'][df.Cabin != 0] = 1 
    df['Embarked'].fillna("S",inplace=True)
    df['Sex'][df.Sex == "male"] = 1
    df['Sex'][df.Sex == "female"] = 0
    df['Age'].fillna(df["Age"].mean(),inplace=True)

    required_data = df.drop([
     'Name',
     'Ticket',
     'Embarked',
     'Pclass',
     #'T_number',
     #'T_group',
     #'Title'
    ],axis="columns")
    
    #One hot encoding
    d1 = pd.get_dummies(df.Pclass)
    d2 = pd.get_dummies(df.Embarked)
    merged_new = pd.concat([required_data,d1],axis="columns")
    
    return merged_new





#Assumption  
# age should be filled by avg
# To calculate mean use imputer class

merged_new = do_preprocessing(df)

#final_set = merged_new.drop(["PassengerId","Survived","SibSp",'Parch','Fare','Cabin','Age'],axis="columns")
final_set = merged_new.drop(["PassengerId","Cabin","SibSp",'Parch',"Survived"],axis="columns")

x= final_set
y = merged_new.Survived
indices = df.index.values 


X_train, X_test,indices_train,indices_test = train_test_split(x,indices, test_size=0.2,random_state=1)
y_train, y_test = y[indices_train],  y[indices_test]


print("Algo: ")
from sklearn.preprocessing import StandardScaler
scaler = StandardScaler()

    
#Random Forest
from sklearn.ensemble import RandomForestClassifier
#model = RandomForestClassifier(n_estimators=30, random_state=42)

model= RandomForestClassifier(bootstrap=True, class_weight=None, criterion='gini',
            max_depth=None, max_features='auto', max_leaf_nodes=None,
            min_impurity_decrease=0.0, min_impurity_split=None,
            min_samples_leaf=1, min_samples_split=2,
            min_weight_fraction_leaf=0.0, n_estimators=300, n_jobs=None,
            oob_score=False, random_state=42, verbose=0, warm_start=False)





model.fit(X_train,y_train)
print("test: ")
print(model.score(X_test,y_test))
print(list(x))
print("train: ")
print(model.score(X_train,y_train))
#applying K-fold cross validation
from sklearn.model_selection import cross_val_score
accuracies = cross_val_score(estimator=model,X=X_train,y=y_train,cv=10)
print("Accuracy mean: ")
print(accuracies.mean())
print("Accuracy Std Dev: ")
print(accuracies.std())
print("F1")
from sklearn.metrics import f1_score
print(f1_score(y_test,model.predict(X_test)))




#### Test & save result


import openpyxl


file="Titanic_report.xlsx"
sheet = "Sheet1"

   
book = openpyxl.load_workbook(path+file)
#print book.get_sheet_names()
# ['Sheet2', 'New Title', 'Sheet1']    # Get a sheet to read
sheet = book.get_sheet_by_name(sheet)    # No of written Rows in sheet
r = sheet.max_row    # No of written Columns in sheet
c = sheet.max_column    # Reading each cell in excel

r = r+1

  
sheet.cell(r, column=2).value = str(model)
sheet.cell(r, column=3).value = str(list(x))
sheet.cell(r, column=4).value = model.score(X_train,y_train)
sheet.cell(r, column=5).value = model.score(X_test,y_test)
sheet.cell(r, column=6).value = accuracies.mean()
sheet.cell(r, column=7).value = f1_score(y_test,model.predict(X_test))
 

#model, Parameter, Scores, Train, test, Accuracy, F1

book.save(path+file)

test_file = path+"test.csv"
df_test = pd.read_csv(test_file)
list(df_test.columns) # for column names
df_test.count() # shows total data 

df_test['Title'] = ""
df_test['T_number'] = df_test['Ticket']
df_test['T_group'] = ""

adding_new_feature(df_test)
merged_new_test = do_preprocessing(df_test)
merged_new_test = merged_new_test.fillna(merged_new_test.mean())
final_set_test = merged_new_test.drop(["PassengerId","Cabin","SibSp",'Parch'],axis="columns")


df_test["Survived"] =  model.predict(final_set_test)

result_download = pd.concat([df_test["PassengerId"],df_test["Survived"]],axis="columns")
result_download.to_csv(path+"result_titanic_rfold_1.csv", sep=',')